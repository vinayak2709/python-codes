import openpyxl
import cleaningTelNum

theFile = openpyxl.load_workbook('Excell_Implenetation2.xlsx')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "Vinayak":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "telephone":
                print("Specific cell on position: {} has value: {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name

def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    print(letter)
    return letter


def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            #take old data and send it to fixing
            telephoneNo = fix_telephone_format(currentSheet[cell_name].value)
            #put new data in cell


            #print(letter + "1")
            if cell_name == (letter + "1"):
                #print(letter + "0")
                #print("aaaaa")
                currentSheet[cell_name].value = "telephone"
            else:
                currentSheet[cell_name].value = telephoneNo



            print("Cell on position: {} has value: {}".format(cell_name, currentSheet[cell_name].value))


#Main Fixning function that calls all other Fix functions
def fix_telephone_format(telephoneNo):
    telephoneNo = cleaningTelNum.remove_first_space_from_tel(telephoneNo)
    telephoneNo = cleaningTelNum.remove_plus_from_tel(telephoneNo)
    telephoneNo = cleaningTelNum.remove_country_code(telephoneNo)
    telephoneNo = cleaningTelNum.place_zero_at_first(telephoneNo)
    telephoneNo = cleaningTelNum.remove_all_characters(telephoneNo)
    return telephoneNo




for sheet in allSheetNames:
    print("\n\nCurrent sheet name is ******* {} \n" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    letter = get_column_letter(specificCellLetter)


    get_all_values_by_cell_letter(letter)

    theFile.save("Customers2.xlsx")
